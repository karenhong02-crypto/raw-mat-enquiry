"""
app.py  –  Flask web front-end for generate_enquiry_raw_mat
Run:  python app.py
Open: http://localhost:5000
"""

import io
import re
import sys
import tempfile
import traceback
import zipfile
from collections import OrderedDict
from copy import copy
from pathlib import Path

from flask import Flask, render_template, request, send_file, jsonify

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.styles.colors import Color
except ImportError:
    sys.exit("Missing dependency. Run:  pip install openpyxl")

# Blue, Accent 5, Lighter 40%  (theme index 8, tint +0.4)
AL6061_FILL = PatternFill(fill_type="solid", fgColor=Color(theme=8, tint=0.3999755859375))

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024   # 32 MB upload limit

ALLOWED_EXT  = {".xlsx", ".xlsm", ".xls"}
DATA_START   = 6
FOOTER1_ROW  = 62
FOOTER2_ROW  = 63
MAT_SUP_PATH = Path(__file__).parent / "Material_Supplier.xlsx"

# Material categories derived from Material & Supplier.xlsx.
# Used as the built-in fallback (and primary source when the xlsx is absent).
# Each entry: name (display), sheet (safe filename stem), keywords (uppercase match tokens).
BUILTIN_CATEGORIES = [
    {
        "name":     "AL5083 or AL6061",
        "sheet":    "AL5083_AL6061",
        "keywords": ["AL5083", "AL6061", "AL"],
    },
    {
        "name":     "MS (bandsaw)",
        "sheet":    "MS",
        "keywords": ["MS"],
    },
    {
        "name":     "Delrin White/Delrin Black/PE (color)/Bakelite (color)/PU/Teflon",
        "sheet":    "Plastic",
        "keywords": ["DELRIN", "PE", "BAKELITE", "PU", "TEFLON", "NYLON"],
    },
    {
        "name":     "SS304",
        "sheet":    "SS304",
        "keywords": ["SS304", "SS"],
    },
    {
        "name":     "S45C",
        "sheet":    "S45C",
        "keywords": ["S45C"],
    },
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def allowed(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXT


def parse_size(s):
    """Parse 'T x W x L' string → (T, W, L) floats."""
    if not s:
        return None, None, None
    s = re.sub(r"[()]", "", str(s))
    parts = re.split(r"\s*x\s*", s.strip(), flags=re.IGNORECASE)
    if len(parts) == 3:
        try:
            return float(parts[0]), float(parts[1]), float(parts[2])
        except ValueError:
            pass
    return None, None, None


def primary_material(m):
    """Normalise material names.
    'AL5083/AL6061' → 'AL5083'
    'MS'            → 'MS (bandsaw)'
    """
    if m and "/" in str(m):
        m = str(m).split("/")[0].strip()
    else:
        m = str(m).strip() if m else ""
    if m.upper().startswith("MS"):
        return f"{m} (bandsaw)"
    return m


def capture_row(ws, row_num, max_col=16):
    saved = {}
    for c in range(1, max_col + 1):
        cell = ws.cell(row_num, c)
        saved[c] = {
            "value":         cell.value,
            "font":          copy(cell.font),
            "fill":          copy(cell.fill),
            "border":        copy(cell.border),
            "alignment":     copy(cell.alignment),
            "number_format": cell.number_format,
        }
    return saved


def restore_row(ws, row_num, saved):
    for c, props in saved.items():
        cell = ws.cell(row_num, c)
        cell.value         = props["value"]
        cell.font          = copy(props["font"])
        cell.fill          = copy(props["fill"])
        cell.border        = copy(props["border"])
        cell.alignment     = copy(props["alignment"])
        cell.number_format = props["number_format"]


def get_pmc_rows(ws_bom):
    rows = []
    for r in range(2, ws_bom.max_row + 1):
        col_o = ws_bom.cell(r, 15).value
        if col_o and str(col_o).strip().upper() == "PMC":
            rows.append({
                "afa":      str(ws_bom.cell(r, 1).value) if ws_bom.cell(r, 1).value else "",
                "qty":      ws_bom.cell(r, 5).value,
                "material": primary_material(ws_bom.cell(r, 6).value),
                "pmc_raw":  ws_bom.cell(r, 14).value,
            })
    return rows


# ---------------------------------------------------------------------------
# Material & Supplier categorisation
# ---------------------------------------------------------------------------

def _parse_categories_from_xlsx(path: str) -> list:
    """Parse Material & Supplier.xlsx into a list of category dicts."""
    wb = load_workbook(path)
    ws = wb.active
    categories = []
    for r in range(2, ws.max_row + 1):
        mat_val = ws.cell(r, 1).value
        if not mat_val:
            continue
        cat_name = str(mat_val).strip()

        sheet_name = re.sub(r"\(.*?\)", "", cat_name)
        sheet_name = re.sub(r"\s*/\s*", "_", sheet_name)
        sheet_name = re.sub(r"\s+or\s+", "_", sheet_name, flags=re.IGNORECASE)
        sheet_name = re.sub(r"[\s_]+", "_", sheet_name).strip("_ ")
        sheet_name = sheet_name[:31]

        raw_parts = re.split(r"\s*/\s*|\s+or\s+", cat_name, flags=re.IGNORECASE)
        keywords = []
        for part in raw_parts:
            kw = re.sub(r"\(.*?\)", "", part).strip().upper()
            if kw:
                keywords.append(kw)

        categories.append({"name": cat_name, "sheet": sheet_name, "keywords": keywords})
    return categories


def get_categories() -> list:
    """Return categories from the bundled xlsx if available, otherwise use built-ins."""
    if MAT_SUP_PATH.exists():
        try:
            return _parse_categories_from_xlsx(str(MAT_SUP_PATH))
        except Exception:
            pass
    return BUILTIN_CATEGORIES


def match_material_category(material: str, categories: list) -> dict | None:
    """Return the first category whose keywords match material, or None."""
    mat_upper = material.upper().strip()
    for cat in categories:
        for kw in cat["keywords"]:
            if mat_upper == kw or mat_upper.startswith(kw):
                return cat
    return None


# Footer text constants — always written regardless of template content
FOOTER1_COL_G  = "OUR REQUIREMENT SIZE"
FOOTER1_COL_L  = "OFFER SIZE (can slightly bigger, CANNOT smaller)"
FOOTER2_COL_L  = "Please highlight In color if size offer different"


def detect_footer_rows(ws) -> tuple:
    """Scan the worksheet bottom-up to find the two footer rows.
    Returns (footer1_row, footer2_row), falling back to hardcoded defaults."""
    f1 = f2 = None
    for r in range(ws.max_row, DATA_START, -1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val is None:
                continue
            s = str(val)
            if f1 is None and "OUR REQUIREMENT SIZE" in s:
                f1 = r
            if f2 is None and "Please highlight" in s:
                f2 = r
        if f1 and f2:
            break
    # If footer rows are adjacent, infer the missing one
    if f1 and not f2:
        f2 = f1 + 1
    if f2 and not f1:
        f1 = f2 - 1
    return (f1 or FOOTER1_ROW), (f2 or FOOTER2_ROW)


# ---------------------------------------------------------------------------
# Sheet filling
# ---------------------------------------------------------------------------

def fill_enquiry_sheet(ws, pmc_rows: list, company: str = "AFA TECHNOLOGIES SDN BHD"):
    """Fill one enquiry worksheet with pmc_rows. Mutates ws in-place."""
    # Update company name in title row, clear project name row
    ws.cell(1, 3).value = f"{company} REQUEST RAW MATERIAL ENQUIRE"
    ws.cell(2, 3).value = None   # blank but formatting preserved

    footer1_row, footer2_row = detect_footer_rows(ws)

    footer1 = capture_row(ws, footer1_row)
    footer2 = capture_row(ws, footer2_row)
    ref_fmt = capture_row(ws, DATA_START)

    n_old = footer1_row - DATA_START
    n_new = len(pmc_rows)

    if n_new > n_old:
        ws.insert_rows(footer1_row, n_new - n_old)
    elif n_new < n_old:
        surplus = n_old - n_new
        ws.delete_rows(footer1_row - surplus, surplus)

    for i, item in enumerate(pmc_rows):
        row = DATA_START + i
        t, w, l = parse_size(item["pmc_raw"])

        ws.cell(row,  1).value = i + 1
        ws.cell(row,  2).value = None
        ws.cell(row,  3).value = item["afa"]
        ws.cell(row,  4).value = item["material"]
        ws.cell(row,  5).value = item["qty"]
        ws.cell(row,  6).value = None
        ws.cell(row,  7).value = item["pmc_raw"]
        ws.cell(row,  8).value = t
        ws.cell(row,  9).value = w
        ws.cell(row, 10).value = l
        ws.cell(row, 11).value = None
        ws.cell(row, 12).value = t
        ws.cell(row, 13).value = w
        ws.cell(row, 14).value = l
        ws.cell(row, 15).value = None
        ws.cell(row, 16).value = f"=E{row}*O{row}"

        for c in range(1, 17):
            dst = ws.cell(row, c)
            dst.font          = copy(ref_fmt[c]["font"])
            dst.fill          = copy(ref_fmt[c]["fill"])
            dst.border        = copy(ref_fmt[c]["border"])
            dst.alignment     = copy(ref_fmt[c]["alignment"])
            dst.number_format = ref_fmt[c]["number_format"]

        # Highlight AL6061 material cell — Blue, Accent 5, Lighter 40%
        if item["material"].upper().startswith("AL6061"):
            ws.cell(row, 4).fill = AL6061_FILL

    last_data_row = DATA_START + n_new - 1
    new_footer1   = last_data_row + 1
    new_footer2   = last_data_row + 2

    restore_row(ws, new_footer1, footer1)
    restore_row(ws, new_footer2, footer2)

    # Always write footer text explicitly so it survives any template variation
    ws.cell(new_footer1,  7).value = FOOTER1_COL_G
    ws.cell(new_footer1, 12).value = FOOTER1_COL_L
    ws.cell(new_footer1, 16).value = f"=SUM(P{DATA_START}:P{last_data_row})"
    ws.cell(new_footer2, 12).value = FOOTER2_COL_L


# ---------------------------------------------------------------------------
# Build functions
# ---------------------------------------------------------------------------

def build_enquiry_bytes(bom_path: str, enq_path: str) -> tuple[bytes, dict]:
    """Single-sheet output (no material segregation)."""
    wb_bom = load_workbook(bom_path)
    wb_enq = load_workbook(enq_path)
    ws_bom = wb_bom.active
    ws_enq = wb_enq.active

    pmc_rows = get_pmc_rows(ws_bom)
    if not pmc_rows:
        raise ValueError("No PMC rows found in BOM file. Check that column O contains 'PMC'.")

    fill_enquiry_sheet(ws_enq, pmc_rows)

    buf = io.BytesIO()
    wb_enq.save(buf)
    buf.seek(0)
    return buf.read(), {"pmc_rows": len(pmc_rows), "sheets": ["All"]}


def build_enquiry_zip(bom_path: str, enq_path: str, company: str = "AFA TECHNOLOGIES SDN BHD") -> tuple[bytes, dict]:
    """Produce a zip containing one .xlsx per material category."""
    wb_bom = load_workbook(bom_path)
    ws_bom = wb_bom.active

    pmc_rows = get_pmc_rows(ws_bom)
    if not pmc_rows:
        raise ValueError("No PMC rows found in BOM file. Check that column O contains 'PMC'.")

    # Sort by material A→Z first, then by AFA code A→Z within each material
    pmc_rows.sort(key=lambda x: (x["material"].upper(), x["afa"].upper()))

    categories = get_categories()

    # Group PMC rows by category
    grouped: OrderedDict = OrderedDict()
    for cat in categories:
        grouped[cat["sheet"]] = {"cat": cat, "rows": []}
    other_rows = []

    for item in pmc_rows:
        cat = match_material_category(item["material"], categories)
        if cat:
            grouped[cat["sheet"]]["rows"].append(item)
        else:
            other_rows.append(item)

    active_groups = [(k, v) for k, v in grouped.items() if v["rows"]]
    if other_rows:
        active_groups.append(("Other", {"cat": {"name": "Other", "sheet": "Other"}, "rows": other_rows}))

    if not active_groups:
        raise ValueError("No PMC rows matched any material category.")

    # Read the template bytes once — reload fresh for every category
    with open(enq_path, "rb") as f:
        enq_bytes = f.read()

    zip_buf = io.BytesIO()
    file_names = []

    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for _, group_data in active_groups:
            cat      = group_data["cat"]
            rows     = group_data["rows"]
            filename = f"Enquiry_{cat['sheet']}.xlsx"

            # Load a fresh copy of the template for each category
            wb_enq = load_workbook(io.BytesIO(enq_bytes))
            fill_enquiry_sheet(wb_enq.active, rows, company)

            xlsx_buf = io.BytesIO()
            wb_enq.save(xlsx_buf)
            zf.writestr(filename, xlsx_buf.getvalue())
            file_names.append(filename)

    zip_buf.seek(0)
    summary = {
        "pmc_rows":   len(pmc_rows),
        "files":      file_names,
        "other_rows": len(other_rows),
    }
    return zip_buf.read(), summary


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/generate", methods=["POST"])
def generate():
    bom_file = request.files.get("bom")
    enq_file = request.files.get("enq")

    # ── Validate uploads ──────────────────────────────────────────────────
    errors = []
    if not bom_file or bom_file.filename == "":
        errors.append("BOM file is required.")
    elif not allowed(bom_file.filename):
        errors.append("BOM file must be an Excel file (.xlsx / .xlsm).")

    if not enq_file or enq_file.filename == "":
        errors.append("Enquiry template file is required.")
    elif not allowed(enq_file.filename):
        errors.append("Enquiry template must be an Excel file (.xlsx / .xlsm).")

    if errors:
        return jsonify({"error": " ".join(errors)}), 400

    # ── Save uploads to temp files ────────────────────────────────────────
    tmp_bom_path = tmp_enq_path = None
    try:
        with (
            tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_bom,
            tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_enq,
        ):
            bom_file.save(tmp_bom.name)
            enq_file.save(tmp_enq.name)
            tmp_bom_path = tmp_bom.name
            tmp_enq_path = tmp_enq.name

        company = request.form.get("company", "AFA TECHNOLOGIES SDN BHD").strip() or "AFA TECHNOLOGIES SDN BHD"
        out_bytes, summary = build_enquiry_zip(tmp_bom_path, tmp_enq_path, company)
        download_name = "Enquiry_Output.zip"
        mimetype      = "application/zip"

    except ValueError as e:
        return jsonify({"error": str(e)}), 422
    except Exception:
        return jsonify({"error": "Processing failed.\n" + traceback.format_exc()}), 500
    finally:
        for p in (tmp_bom_path, tmp_enq_path):
            if p:
                Path(p).unlink(missing_ok=True)

    # ── Stream the file back ──────────────────────────────────────────────
    buf = io.BytesIO(out_bytes)
    buf.seek(0)
    response = send_file(
        buf,
        as_attachment=True,
        download_name=download_name,
        mimetype=mimetype,
    )
    files_made = summary.get("files") or summary.get("sheets", [])
    response.headers["X-Files-Generated"] = ", ".join(files_made)
    response.headers["X-PMC-Rows"]        = str(summary.get("pmc_rows", 0))
    return response


# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=5000)
